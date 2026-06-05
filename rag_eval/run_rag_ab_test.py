#!/usr/bin/env python3
"""Build and optionally run a RAG A/B evaluation workbook.

The script first consolidates historical no-RAG evaluation cases from
evaluation/reports and evaluation/data, then optionally runs the current
RAG-enabled backend on the same unique case set.
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import math
import sys
import time
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REPORTS_DIR = ROOT / "evaluation" / "reports"
DEFAULT_DATA_DIR = ROOT / "evaluation" / "data"
DEFAULT_BACKEND_URL = "http://localhost:8081"
DEFAULT_PRE_INQUIRY_PATH = "/api/agent/pre-consultation"
RAG_HEALTH_URL = "http://localhost:18080/health"
RAG_RETRIEVE_URL = "http://localhost:18080/rag/retrieve"
PROGRESS_FILE = "rag_ab_test_progress.json"

CASE_FIELDS = [
    "case_id",
    "case_title",
    "test_group",
    "population_type",
    "disease_category",
    "mode",
    "user_input",
    "expected_department",
    "expected_urgency",
    "must_ask_points",
    "red_flags",
    "forbidden_actions",
    "expected_response_points",
    "manual_review_level",
    "scoring_focus",
    "source_batch",
    "duplicate_from",
    "no_rag_score_missing",
]

SCORE_FIELDS = [
    "medical_safety_score",
    "department_accuracy_score",
    "inquiry_completeness_score",
    "mode_compliance_score",
    "anti_misleading_score",
    "expression_score",
    "total_score",
    "is_one_vote_veto",
    "is_suspected_failure",
    "request_error",
    "manual_review_required",
    "must_ask_hit_count",
    "must_ask_total_count",
    "must_ask_hit_rate",
]

GROUP_LABELS = {
    "quick_common": "快速问诊普通",
    "quick_red_flag": "红旗风险",
    "deep_structured": "深度问诊",
    "special_population_child": "儿童专项",
    "special_population_elderly": "老年人专项",
    "special_population_pregnancy": "孕产妇专项",
    "chronic_medication": "慢病用药专项",
    "immunocompromised_postop": "免疫低下/肿瘤/术后专项",
    "mental_health_self_harm": "精神心理专项",
    "disease_category_extension": "其他高风险疾病扩展",
}

POPULATION_HINTS = {
    "儿童": ["孩子", "宝宝", "婴儿", "岁孩子", "儿童"],
    "老人": ["老人", "老年", "我爸", "我妈", "爷爷", "奶奶", "72岁", "68岁", "老人"],
    "孕产妇": ["怀孕", "孕", "产后", "胎动", "恶露"],
    "慢病患者": ["糖尿病", "高血压", "冠心病", "肾功能", "哮喘", "抗凝", "激素"],
    "免疫低下": ["化疗", "免疫抑制", "移植", "肿瘤", "术后"],
    "精神心理风险": ["不想活", "自杀", "绝望", "幻听", "被害", "焦虑"],
}


def load_module(name: str, path: Path) -> Any:
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


RUN_EVAL = load_module("run_pre_consultation_eval", ROOT / "evaluation" / "scripts" / "run_pre_consultation_eval.py")
BUILD_SCORE = load_module("build_extended_eval_report_excel", ROOT / "evaluation" / "scripts" / "build_extended_eval_report_excel.py")


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_cases_from_jsonl(path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if line:
                cases.append(json.loads(line))
    return cases


def latest_user_input(case: dict[str, Any]) -> str:
    turns = case.get("turns") or []
    user_turns = [str(turn.get("content") or "") for turn in turns if turn.get("role") == "user"]
    return "\n".join(user_turns)


def first_user_input(case: dict[str, Any]) -> str:
    turns = case.get("turns") or []
    for turn in turns:
        if turn.get("role") == "user" and turn.get("content"):
            return str(turn["content"])
    return ""


def infer_population(case: dict[str, Any]) -> str:
    profile = case.get("patientProfile") or {}
    special = str(profile.get("specialPopulation") or "").strip()
    if special:
        return special
    text = f"{case.get('title', '')} {latest_user_input(case)}"
    for label, hints in POPULATION_HINTS.items():
        if any(hint in text for hint in hints):
            return label
    return "普通成人"


def infer_disease_category(case: dict[str, Any]) -> str:
    text = f"{case.get('title', '')} {latest_user_input(case)}"
    mapping = [
        ("心脑血管", ["胸", "心", "卒中", "偏瘫", "晕厥", "血压"]),
        ("呼吸系统", ["咳", "喘", "呼吸", "咯血", "肺"]),
        ("消化系统", ["腹", "吐", "泻", "黑便", "呕血"]),
        ("泌尿生殖", ["尿", "肾", "睾丸", "孕", "胎"]),
        ("内分泌", ["糖尿病", "低血糖", "高血糖", "甲亢"]),
        ("皮肤感染", ["皮疹", "红肿", "蜂窝", "带状疱疹", "紫癜"]),
        ("眼耳鼻喉", ["眼", "鼻", "耳", "咽", "听力", "视力"]),
        ("外伤中毒", ["外伤", "摔", "误服", "咬伤", "烫伤", "骨折"]),
        ("精神心理", ["不想活", "焦虑", "失眠", "幻听", "妄想"]),
    ]
    for label, hints in mapping:
        if any(hint in text for hint in hints):
            return label
    return str(case.get("caseType") or "")


def classify_source(path: Path) -> str:
    name = path.name.lower()
    if "extended_refined" in name or "extended_special" in name:
        return "supplemental_category"
    if "223123" in name or "failed" in name:
        return "failed_21_retest"
    if "222206" in name:
        return "initial_50"
    return "initial_50"


def canonical_case(case: dict[str, Any], source_batch: str) -> dict[str, Any]:
    expected = case.get("expected") or case.get("expectedRecord") or {}
    return {
        "case_id": str(case.get("caseId") or ""),
        "case_title": str(case.get("title") or ""),
        "test_group": GROUP_LABELS.get(str(case.get("caseType") or ""), str(case.get("caseType") or "")),
        "population_type": infer_population(case),
        "disease_category": infer_disease_category(case),
        "mode": str(case.get("mode") or "quick"),
        "mode_defaulted": not bool(case.get("mode")),
        "user_input": latest_user_input(case),
        "first_user_input": first_user_input(case),
        "expected_department": stringify(expected.get("recommendedDepartments") or expected.get("departments") or []),
        "expected_urgency": str(expected.get("urgency") or ""),
        "must_ask_points": stringify(expected.get("mustAskOrMention") or []),
        "red_flags": stringify(expected.get("redFlagsToMention") or []),
        "forbidden_actions": stringify(expected.get("mustNotSay") or expected.get("mustNotInvent") or case.get("hardFailRules") or []),
        "expected_response_points": stringify(expected),
        "manual_review_level": "建议复核",
        "scoring_focus": stringify(case.get("scoring") or {}),
        "source_batch": source_batch,
        "duplicate_from": "",
        "no_rag_score_missing": True,
        "_raw_case": case,
    }


def completeness_score(case_row: dict[str, Any]) -> int:
    keys = ["user_input", "expected_department", "expected_urgency", "must_ask_points", "forbidden_actions"]
    return sum(1 for key in keys if case_row.get(key)) + len(str(case_row.get("expected_response_points") or ""))


def dedupe_cases(case_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_key: dict[str, dict[str, Any]] = {}
    duplicates: list[dict[str, Any]] = []
    for row in case_rows:
        key = row["case_id"] or "|".join([
            normalize_text(row.get("user_input")),
            normalize_text(row.get("test_group")),
            normalize_text(row.get("case_title")),
            normalize_text(row.get("expected_department")),
            normalize_text(row.get("expected_urgency")),
        ])
        text_key = normalize_text(row.get("user_input"))
        existing_key = None
        if key in by_key:
            existing_key = key
        elif text_key:
            for candidate_key, candidate in by_key.items():
                if normalize_text(candidate.get("user_input")) == text_key:
                    existing_key = candidate_key
                    break
        if existing_key is None:
            by_key[key] = row
            continue
        existing = by_key[existing_key]
        duplicates.append({
            "kept_case_id": existing.get("case_id"),
            "duplicate_case_id": row.get("case_id"),
            "kept_source": existing.get("source_batch"),
            "duplicate_source": row.get("source_batch"),
            "duplicate_title": row.get("case_title"),
        })
        sources = set(str(existing.get("source_batch") or "").split("+")) | set(str(row.get("source_batch") or "").split("+"))
        sources.discard("")
        existing["source_batch"] = "multiple_sources" if len(sources) > 1 else (next(iter(sources)) if sources else "")
        dup_from = [item for item in [existing.get("duplicate_from"), row.get("case_id"), row.get("source_batch")] if item]
        existing["duplicate_from"] = "; ".join(dict.fromkeys(map(str, dup_from)))
        if completeness_score(row) > completeness_score(existing):
            row["source_batch"] = existing["source_batch"]
            row["duplicate_from"] = existing["duplicate_from"]
            by_key[existing_key] = row
    return list(by_key.values()), duplicates


def score_response(case: dict[str, Any], response: Any, dry_run: bool = False) -> dict[str, Any]:
    result = RUN_EVAL.evaluate_response(case, response, dry_run)
    score = BUILD_SCORE.score_result(result)
    must_total = int(result.get("mustAskTotal") or 0)
    must_hit = int(result.get("mustAskHitCount") or 0)
    request_error = BUILD_SCORE.request_error(result.get("rawResponse"))
    return {
        "medical_safety_score": score["medicalSafety"],
        "department_accuracy_score": score["departmentAccuracy"],
        "inquiry_completeness_score": score["consultationCompleteness"],
        "mode_compliance_score": score["modeCompliance"],
        "anti_misleading_score": score["antiMisleading"],
        "expression_score": score["userExperience"],
        "total_score": score["total"],
        "is_one_vote_veto": bool(score["oneVoteFail"]),
        "is_suspected_failure": bool(result.get("suspectedFail")),
        "request_error": bool(request_error),
        "request_error_message": request_error,
        "manual_review_required": bool(result.get("manualReviewRequired")),
        "must_ask_hit_count": must_hit,
        "must_ask_total_count": must_total,
        "must_ask_hit_rate": round(must_hit / must_total, 4) if must_total else None,
        "model_response": BUILD_SCORE.final_reply(result.get("rawResponse")),
        "score_reason": stringify({
            "department_hits": result.get("departmentHits"),
            "urgency_hits": result.get("urgencyHits"),
            "must_not_hits": result.get("mustNotHits"),
            "hard_fail_hits": result.get("hardFailHits"),
        }),
        "_eval_result": result,
    }


def normalize_historical_result(result: dict[str, Any]) -> dict[str, Any]:
    score = BUILD_SCORE.score_result(result)
    must_total = int(result.get("mustAskTotal") or 0)
    must_hit = int(result.get("mustAskHitCount") or 0)
    request_error = BUILD_SCORE.request_error(result.get("rawResponse"))
    return {
        "case_id": str(result.get("caseId") or ""),
        "case_title": str(result.get("title") or ""),
        "test_group": GROUP_LABELS.get(str(result.get("caseType") or ""), str(result.get("caseType") or "")),
        "mode": str(result.get("mode") or "quick"),
        "medical_safety_score": score["medicalSafety"],
        "department_accuracy_score": score["departmentAccuracy"],
        "inquiry_completeness_score": score["consultationCompleteness"],
        "mode_compliance_score": score["modeCompliance"],
        "anti_misleading_score": score["antiMisleading"],
        "expression_score": score["userExperience"],
        "total_score": score["total"],
        "is_one_vote_veto": bool(score["oneVoteFail"]),
        "is_suspected_failure": bool(result.get("suspectedFail")),
        "request_error": bool(request_error),
        "request_error_message": request_error,
        "manual_review_required": bool(result.get("manualReviewRequired")),
        "must_ask_hit_count": must_hit,
        "must_ask_total_count": must_total,
        "must_ask_hit_rate": round(must_hit / must_total, 4) if must_total else None,
        "model_response": BUILD_SCORE.final_reply(result.get("rawResponse")),
        "score_reason": stringify({
            "department_hits": result.get("departmentHits"),
            "urgency_hits": result.get("urgencyHits"),
            "must_not_hits": result.get("mustNotHits"),
            "hard_fail_hits": result.get("hardFailHits"),
        }),
    }


def discover_history(reports_dir: Path, data_dir: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    if not reports_dir.exists():
        raise FileNotFoundError(f"reports-dir does not exist: {reports_dir}")
    case_rows: list[dict[str, Any]] = []
    no_rag_scores: dict[str, dict[str, Any]] = {}
    history_files: list[dict[str, Any]] = []

    for case_file in sorted(data_dir.glob("*.jsonl")):
        source = classify_source(case_file)
        cases = read_cases_from_jsonl(case_file)
        for case in cases:
            case_rows.append(canonical_case(case, source))
        history_files.append({
            "path": str(case_file),
            "file_type": "jsonl case source",
            "sheets": "",
            "rows": len(cases),
            "columns": "caseId, caseType, mode, title, turns, expected, scoring",
            "purpose": f"测试样例源文件/{source}",
        })

    def is_generated_ab_file(path: Path) -> bool:
        return path.name.startswith("medical_rag_ab_test_")

    for json_file in sorted(reports_dir.glob("eval_report*.json")):
        if is_generated_ab_file(json_file):
            continue
        data = read_json(json_file)
        if not isinstance(data, list):
            continue
        source = classify_source(json_file)
        history_files.append({
            "path": str(json_file),
            "file_type": "json result",
            "sheets": "(json-list)",
            "rows": len(data),
            "columns": ", ".join(list(data[0].keys())[:20]) if data else "",
            "purpose": f"无 RAG 历史评分/{source}",
        })
        for result in data:
            case_id = str(result.get("caseId") or "")
            if not case_id:
                continue
            score = normalize_historical_result(result)
            score["source_batch"] = source
            score["test_time"] = datetime.fromtimestamp(json_file.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            score["_mtime"] = json_file.stat().st_mtime
            previous = no_rag_scores.get(case_id)
            if previous is None or score["_mtime"] >= previous.get("_mtime", 0):
                no_rag_scores[case_id] = score

    for xlsx_file in sorted(reports_dir.glob("*.xlsx")):
        if is_generated_ab_file(xlsx_file):
            continue
        history_files.append({
            "path": str(xlsx_file),
            "file_type": "xlsx report",
            "sheets": "Excel 工作簿，结构由历史报告保留",
            "rows": "",
            "columns": "",
            "purpose": "人工查看/历史评分汇总",
        })
    for text_file in sorted(list(reports_dir.glob("*.md")) + list(reports_dir.glob("*.txt")) + list(reports_dir.glob("*.csv"))):
        if is_generated_ab_file(text_file):
            continue
        history_files.append({
            "path": str(text_file),
            "file_type": text_file.suffix.lower().lstrip("."),
            "sheets": "",
            "rows": "",
            "columns": "",
            "purpose": "历史分析/CSV摘要",
        })
    return case_rows, no_rag_scores, history_files


def check_url(url: str, timeout: float = 5.0) -> tuple[bool, str]:
    try:
        with urllib.request.urlopen(url, timeout=timeout) as response:
            return 200 <= response.status < 500, response.read(300).decode("utf-8", errors="ignore")
    except Exception as exc:
        return False, str(exc)


def check_tcp(host: str, port: int, timeout: float = 3.0) -> tuple[bool, str]:
    import socket

    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True, "tcp ok"
    except Exception as exc:
        return False, str(exc)


def post_json(url: str, payload: dict[str, Any], timeout: float) -> dict[str, Any]:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def check_rag_retrieval() -> dict[str, Any]:
    ok, body = check_url(RAG_HEALTH_URL)
    if not ok:
        return {"rag_service_accessible": False, "error": body}
    payload = {"query": "胸痛大汗需要急诊吗", "top_k": 3, "scene": "pre_inquiry"}
    try:
        response = post_json(RAG_RETRIEVE_URL, payload, timeout=30)
        chunks = response.get("chunks") or []
        return {
            "rag_service_accessible": True,
            "rag_retrieve_success": bool(response.get("success")),
            "rag_hit_count": len(chunks),
            "rag_titles": "; ".join(str(chunk.get("title") or "") for chunk in chunks),
        }
    except Exception as exc:
        return {"rag_service_accessible": True, "rag_retrieve_success": False, "error": str(exc)}


def check_milvus() -> dict[str, Any]:
    try:
        from pymilvus import Collection, connections, utility
    except Exception as exc:
        return {"milvus_accessible": False, "error": f"pymilvus import failed: {exc}"}
    try:
        connections.connect(alias="rag_ab_eval", host="localhost", port="19530")
        exists = utility.has_collection("medical_rag_chunks", using="rag_ab_eval")
        if not exists:
            return {"milvus_accessible": True, "collection_exists": False, "num_entities": 0}
        collection = Collection("medical_rag_chunks", using="rag_ab_eval")
        collection.load()
        rows = collection.query(expr='doc_type != ""', output_fields=["doc_type"], limit=10000)
        counts = Counter(row.get("doc_type") for row in rows)
        return {
            "milvus_accessible": True,
            "collection_exists": True,
            "num_entities": collection.num_entities,
            "doc_type_counts": dict(counts),
        }
    except Exception as exc:
        return {"milvus_accessible": False, "error": str(exc)}


def read_rag_enabled() -> str:
    config = ROOT / "agent-server" / "src" / "main" / "resources" / "application.yml"
    if not config.exists():
        return "unknown"
    text = config.read_text(encoding="utf-8", errors="ignore")
    if "rag:" in text and "enabled:" in text:
        return "configured true by default/env" if "${RAG_ENABLED:true}" in text else "configured, please verify env"
    return "not found"


def run_rag_case(base_url: str, case: dict[str, Any], timeout: float, multi_turn: bool) -> tuple[Any, float]:
    start = time.perf_counter()
    if case.get("caseType") == "record_generation":
        response = RUN_EVAL.post_json(
            base_url.rstrip("/") + "/api/agent/medical-record-drafts/generate",
            RUN_EVAL.record_payload(case),
            timeout=timeout,
        )
    else:
        response = RUN_EVAL.run_pre_consultation_case(base_url, case, timeout=timeout, multi_turn=multi_turn)
    elapsed = (time.perf_counter() - start) * 1000
    return response, elapsed


def is_service_unavailable_error(exc: BaseException) -> bool:
    text = str(exc)
    return any(marker in text for marker in [
        "WinError 10061",
        "Connection refused",
        "actively refused",
        "无法连接",
        "timed out",
        "Read timed out",
    ])


def retrieve_rag_details(query: str, mode: str) -> dict[str, Any]:
    scene = "deep_inquiry" if mode == "deep" else "pre_inquiry"
    payload = {"query": query, "top_k": 8, "scene": scene}
    start = time.perf_counter()
    try:
        response = post_json(RAG_RETRIEVE_URL, payload, timeout=30)
        elapsed = round((time.perf_counter() - start) * 1000, 1)
        chunks = response.get("chunks") or []
        return {
            "query": query,
            "rag_hit_count": len(chunks),
            "rag_doc_types": "; ".join(str(chunk.get("doc_type") or "") for chunk in chunks),
            "rag_titles": "; ".join(str(chunk.get("title") or "") for chunk in chunks),
            "rag_scores": "; ".join(str(round(float(chunk.get("score") or 0), 4)) for chunk in chunks),
            "rag_context_chars": sum(len(str(chunk.get("chunk_text") or "")) for chunk in chunks),
            "rag_request_time_ms": elapsed,
            "rag_failed_but_fallback": False,
        }
    except Exception as exc:
        return {
            "query": query,
            "rag_hit_count": 0,
            "rag_doc_types": "",
            "rag_titles": "",
            "rag_scores": "",
            "rag_context_chars": 0,
            "rag_request_time_ms": None,
            "rag_failed_but_fallback": True,
            "rag_error": str(exc),
        }


def mean(values: list[Any]) -> float | None:
    nums = [float(value) for value in values if value is not None and value != "" and not (isinstance(value, float) and math.isnan(value))]
    return round(sum(nums) / len(nums), 4) if nums else None


def add_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], headers: list[str]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, col in enumerate(sheet.columns, 1):
        max_len = max(len(str(cell.value or "")) for cell in col[:80])
        sheet.column_dimensions[get_column_letter(index)].width = max(10, min(max_len + 2, 55))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_report_md(path: Path, summary: dict[str, Any], group_rows: list[dict[str, Any]], notes: list[str]) -> None:
    lines = [
        "# Medical RAG A/B Test Report",
        "",
        "## 样例整理结果",
        f"- 历史测试文件数量：{summary.get('history_file_count')}",
        f"- 去重前 case 数：{summary.get('case_count_before_dedupe')}",
        f"- 去重后唯一 case 数：{summary.get('case_count_after_dedupe')}",
        f"- 重复 case 数：{summary.get('duplicate_count')}",
        f"- 无历史无 RAG 分数 case 数：{summary.get('no_rag_score_missing_count')}",
        "",
        "## RAG 测试环境",
        f"- rag.enabled：{summary.get('rag_enabled_status')}",
        f"- RAG 服务可访问：{summary.get('rag_service_accessible')}",
        f"- Milvus collection 存在：{summary.get('collection_exists')}",
        f"- num_entities：{summary.get('num_entities')}",
        f"- doc_type counts：{summary.get('doc_type_counts')}",
        "",
        "## 总体对比",
        f"- 无 RAG 平均总分：{summary.get('no_rag_avg_total_score')}",
        f"- 有 RAG 平均总分：{summary.get('rag_avg_total_score')}",
        f"- 平均分变化：{summary.get('avg_score_delta')}",
        f"- 无 RAG 必问命中率：{summary.get('no_rag_avg_must_ask_hit_rate')}",
        f"- 有 RAG 必问命中率：{summary.get('rag_avg_must_ask_hit_rate')}",
        f"- 必问命中率变化：{summary.get('avg_must_ask_hit_rate_delta')}",
        f"- 改善 case 数：{summary.get('improved_count')}",
        f"- 退步 case 数：{summary.get('regressed_count')}",
        f"- 持平 case 数：{summary.get('unchanged_count')}",
        f"- 跳过已完成 case 数：{summary.get('skipped_completed_count')}",
        f"- 本次新执行 case 数：{summary.get('newly_executed_count')}",
        f"- 重试失败 case 数：{summary.get('retried_failed_count')}",
        "",
        "## 分组结论",
    ]
    for row in group_rows:
        lines.append(
            f"- {row.get('test_group')}: cases={row.get('case_count')}, "
            f"no_rag_avg={row.get('no_rag_avg_total_score')}, "
            f"rag_avg={row.get('rag_avg_total_score')}, delta={row.get('avg_score_delta')}"
        )
    lines.extend(["", "## 注意事项"])
    lines.extend(f"- {note}" for note in notes)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "是"}


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header or "") for header in rows[0]]
    result = []
    for values in rows[1:]:
        result.append({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    workbook.close()
    return result


def latest_ab_workbook(output_dir: Path) -> Path | None:
    files = sorted(output_dir.glob("medical_rag_ab_test_results_*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    return files[0] if files else None


def cache_entry_from_rows(rag_row: dict[str, Any], retrieval_row: dict[str, Any] | None = None) -> dict[str, Any]:
    retrieval_row = retrieval_row or {}
    case_id = str(rag_row.get("case_id") or "")
    return {
        "case_id": case_id,
        "user_input": rag_row.get("user_input", ""),
        "rag_model_response": rag_row.get("rag_model_response", ""),
        "rag_total_score": rag_row.get("total_score", ""),
        "rag_must_ask_hit_count": rag_row.get("must_ask_hit_count", ""),
        "rag_must_ask_total_count": rag_row.get("must_ask_total_count", ""),
        "rag_must_ask_hit_rate": rag_row.get("must_ask_hit_rate", ""),
        "rag_request_error": truthy(rag_row.get("request_error")),
        "rag_error_message": rag_row.get("request_error_message", "") or rag_row.get("rag_score_reason", ""),
        "rag_hit_count": retrieval_row.get("rag_hit_count", rag_row.get("rag_hit_count", "")),
        "rag_doc_types": retrieval_row.get("rag_doc_types", rag_row.get("rag_doc_types", "")),
        "rag_titles": retrieval_row.get("rag_titles", rag_row.get("rag_titles", "")),
        "rag_scores": retrieval_row.get("rag_scores", ""),
        "rag_context_chars": retrieval_row.get("rag_context_chars", ""),
        "request_time_ms": rag_row.get("rag_request_time_ms", retrieval_row.get("rag_request_time_ms", "")),
        "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rag_score": rag_row,
        "retrieval": retrieval_row,
    }


def load_resume_cache(output_dir: Path) -> tuple[dict[str, dict[str, Any]], str]:
    progress_path = output_dir / PROGRESS_FILE
    if progress_path.exists():
        data = read_json(progress_path)
        entries = data.get("entries", data) if isinstance(data, dict) else data
        if isinstance(entries, list):
            return {str(entry.get("case_id")): entry for entry in entries if entry.get("case_id")}, str(progress_path)
        if isinstance(entries, dict):
            return {str(key): value for key, value in entries.items()}, str(progress_path)

    workbook_path = latest_ab_workbook(output_dir)
    if not workbook_path:
        return {}, ""
    rag_rows = sheet_rows(workbook_path, "rag_scores")
    retrieval_rows = {str(row.get("case_id") or ""): row for row in sheet_rows(workbook_path, "rag_retrieval_details")}
    cache: dict[str, dict[str, Any]] = {}
    for row in rag_rows:
        case_id = str(row.get("case_id") or "")
        if not case_id:
            continue
        has_response = bool(str(row.get("rag_model_response") or "").strip())
        has_error = truthy(row.get("request_error"))
        has_score = row.get("total_score") not in (None, "")
        if has_response or has_error or has_score:
            cache[case_id] = cache_entry_from_rows(row, retrieval_rows.get(case_id))
    return cache, str(workbook_path)


def save_progress(output_dir: Path, cache: dict[str, dict[str, Any]], meta: dict[str, Any]) -> Path:
    progress_path = output_dir / PROGRESS_FILE
    payload = {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "meta": meta,
        "entries": list(cache.values()),
    }
    progress_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return progress_path


def score_from_cache_entry(entry: dict[str, Any]) -> dict[str, Any]:
    row = dict(entry.get("rag_score") or {})
    if not row:
        row = {
            "total_score": entry.get("rag_total_score", ""),
            "must_ask_hit_count": entry.get("rag_must_ask_hit_count", ""),
            "must_ask_total_count": entry.get("rag_must_ask_total_count", ""),
            "must_ask_hit_rate": entry.get("rag_must_ask_hit_rate", ""),
            "request_error": entry.get("rag_request_error", ""),
            "request_error_message": entry.get("rag_error_message", ""),
            "rag_model_response": entry.get("rag_model_response", ""),
            "rag_score_reason": entry.get("rag_error_message", ""),
        }
    row["model_response"] = row.get("model_response", row.get("rag_model_response", ""))
    row["score_reason"] = row.get("score_reason", row.get("rag_score_reason", ""))
    return row


def build_summaries(comparison: list[dict[str, Any]], group_key: str) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in comparison:
        grouped[str(row.get(group_key) or "未分类")].append(row)
    rows = []
    for key, items in sorted(grouped.items()):
        rows.append({
            group_key: key,
            "case_count": len(items),
            "no_rag_avg_total_score": mean([item.get("no_rag_total_score") for item in items]),
            "rag_avg_total_score": mean([item.get("rag_total_score") for item in items]),
            "avg_score_delta": mean([item.get("score_delta") for item in items]),
            "no_rag_avg_must_ask_hit_rate": mean([item.get("no_rag_must_ask_hit_rate") for item in items]),
            "rag_avg_must_ask_hit_rate": mean([item.get("rag_must_ask_hit_rate") for item in items]),
            "avg_must_ask_hit_rate_delta": mean([item.get("must_ask_hit_rate_delta") for item in items]),
            "no_rag_failure_count": sum(1 for item in items if item.get("no_rag_is_suspected_failure")),
            "rag_failure_count": sum(1 for item in items if item.get("rag_is_suspected_failure")),
            "improved_count": sum(1 for item in items if item.get("improved")),
            "regressed_count": sum(1 for item in items if item.get("regressed")),
            "unchanged_count": sum(1 for item in items if item.get("unchanged")),
        })
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--reports-dir", default=str(DEFAULT_REPORTS_DIR))
    parser.add_argument("--output-dir")
    parser.add_argument("--output")
    parser.add_argument("--rag-enabled", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--no-rag-only", action="store_true")
    parser.add_argument("--limit", type=int)
    parser.add_argument("--case-id", action="append", help="Only include the specified case_id. Can be repeated.")
    parser.add_argument("--resume", action="store_true")
    parser.add_argument("--sleep-ms", type=int, default=0)
    parser.add_argument("--backend-url", default=DEFAULT_BACKEND_URL)
    parser.add_argument("--pre-inquiry-path", default=DEFAULT_PRE_INQUIRY_PATH)
    parser.add_argument("--mode", default="quick")
    parser.add_argument("--request-timeout", type=float, default=180.0)
    parser.add_argument("--multi-turn", action="store_true", default=True)
    parser.add_argument("--rescore-no-rag", action="store_true")
    args = parser.parse_args()

    reports_dir = Path(args.reports_dir)
    output_dir = Path(args.output_dir) if args.output_dir else reports_dir
    if not reports_dir.exists():
        print(f"reports-dir 不存在，请确认路径：{reports_dir}", file=sys.stderr)
        return 2
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = Path(args.output) if args.output else output_dir / f"medical_rag_ab_test_results_{timestamp}.xlsx"
    output_md = output_dir / f"medical_rag_ab_test_report_{timestamp}.md"

    case_rows, no_rag_scores, history_files = discover_history(reports_dir, DEFAULT_DATA_DIR)
    unique_cases, duplicates = dedupe_cases(case_rows)
    if args.limit:
        unique_cases = unique_cases[: args.limit]
    if args.case_id:
        wanted_case_ids = set(args.case_id)
        unique_cases = [row for row in unique_cases if row["case_id"] in wanted_case_ids]

    for row in unique_cases:
        row["no_rag_score_missing"] = row["case_id"] not in no_rag_scores

    env = {
        "rag_enabled_status": read_rag_enabled(),
        "rag_service_accessible": None,
        "collection_exists": None,
        "num_entities": None,
        "doc_type_counts": None,
    }
    notes: list[str] = []
    if args.no_rag_only or not args.rag_enabled:
        notes.append("本次以 no-rag-only 运行：只整理历史无 RAG 样例和分数，不调用当前 RAG 后端。")
    else:
        backend_host = args.backend_url.replace("http://", "").replace("https://", "").split("/")[0].split(":")[0] or "localhost"
        backend_port = 443 if args.backend_url.startswith("https://") else 80
        if ":" in args.backend_url.replace("http://", "").replace("https://", "").split("/")[0]:
            backend_port = int(args.backend_url.replace("http://", "").replace("https://", "").split("/")[0].split(":")[1])
        backend_ok, backend_body = check_tcp(backend_host, backend_port)
        if not backend_ok:
            print(f"后端服务不可访问，停止正式 RAG 测试：{backend_body}", file=sys.stderr)
            return 3
        rag_status = check_rag_retrieval()
        env.update(rag_status)
        if not rag_status.get("rag_service_accessible") or not rag_status.get("rag_retrieve_success"):
            print(f"RAG 服务不可用，停止正式 RAG 测试：{rag_status}", file=sys.stderr)
            return 4
        milvus_status = check_milvus()
        env.update(milvus_status)
        if not milvus_status.get("milvus_accessible") or not milvus_status.get("collection_exists") or not milvus_status.get("num_entities"):
            print(f"Milvus collection 不可用，停止正式 RAG 测试：{milvus_status}", file=sys.stderr)
            return 5

    no_rag_rows: list[dict[str, Any]] = []
    rag_rows: list[dict[str, Any]] = []
    comparison_rows: list[dict[str, Any]] = []
    retrieval_rows: list[dict[str, Any]] = []
    resume_cache: dict[str, dict[str, Any]] = {}
    resume_source = ""
    skipped_completed_count = 0
    newly_executed_count = 0
    retried_failed_count = 0
    if args.resume:
        resume_cache, resume_source = load_resume_cache(output_dir)
        if resume_source:
            notes.append(f"断点续跑已读取缓存来源：{resume_source}")
        else:
            notes.append("断点续跑未找到已有缓存或 A/B 工作簿，将从当前唯一 case 集开始执行。")
    consecutive_service_failures = 0
    aborted_due_to_service_failure = False

    for index, case_row in enumerate(unique_cases, 1):
        case = case_row["_raw_case"]
        no_rag = no_rag_scores.get(case_row["case_id"], {})
        no_rag_rows.append({
            "case_id": case_row["case_id"],
            "case_title": case_row["case_title"],
            "test_group": case_row["test_group"],
            "user_input": case_row["user_input"],
            **{field: no_rag.get(field, "") for field in SCORE_FIELDS},
            "no_rag_model_response": no_rag.get("model_response", ""),
            "no_rag_score_reason": no_rag.get("score_reason", ""),
            "source_batch": no_rag.get("source_batch", case_row.get("source_batch", "")),
            "test_time": no_rag.get("test_time", ""),
        })

        rag_score: dict[str, Any] = {}
        retrieval = {
            "case_id": case_row["case_id"],
            "query": case_row["first_user_input"] or case_row["user_input"],
            "rag_hit_count": "",
            "rag_doc_types": "",
            "rag_titles": "",
            "rag_scores": "",
            "rag_context_chars": "",
            "rag_request_time_ms": "",
            "rag_failed_but_fallback": "",
        }
        cached_entry = resume_cache.get(case_row["case_id"]) if args.resume else None
        cached_success = bool(
            cached_entry
            and not cached_entry.get("rag_request_error")
            and cached_entry.get("rag_total_score") not in (None, "")
        )
        if cached_success:
            skipped_completed_count += 1
            rag_score = score_from_cache_entry(cached_entry)
            retrieval.update(cached_entry.get("retrieval") or {})
        elif not args.no_rag_only and args.rag_enabled:
            if cached_entry and cached_entry.get("rag_request_error"):
                retried_failed_count += 1
            print(f"[{index}/{len(unique_cases)}] RAG testing {case_row['case_id']} {case_row['case_title']}")
            query = case_row["first_user_input"] or case_row["user_input"]
            retrieval.update(retrieve_rag_details(query, case_row.get("mode") or args.mode))
            try:
                response, elapsed_ms = run_rag_case(args.backend_url, case, timeout=args.request_timeout, multi_turn=args.multi_turn)
                rag_score = score_response(case, response, dry_run=False)
                rag_score["rag_request_time_ms"] = round(elapsed_ms, 1)
                rag_score["rag_model_response"] = rag_score.get("model_response", "")
                rag_score["rag_score_reason"] = rag_score.get("score_reason", "")
                newly_executed_count += 1
                consecutive_service_failures = 0
            except (TimeoutError, urllib.error.URLError, json.JSONDecodeError) as exc:
                response = {"error": f"Service request failed: {exc}"}
                rag_score = score_response(case, response, dry_run=False)
                rag_score["rag_model_response"] = ""
                rag_score["rag_score_reason"] = str(exc)
                newly_executed_count += 1
                if is_service_unavailable_error(exc):
                    consecutive_service_failures += 1
                else:
                    consecutive_service_failures = 0
            if args.resume:
                progress_rag_row = {
                    "case_id": case_row["case_id"],
                    "user_input": case_row["user_input"],
                    **{field: rag_score.get(field, "") for field in SCORE_FIELDS},
                    "request_error_message": rag_score.get("request_error_message", ""),
                    "rag_request_time_ms": rag_score.get("rag_request_time_ms", retrieval.get("rag_request_time_ms", "")),
                    "rag_hit_count": retrieval.get("rag_hit_count", ""),
                    "rag_doc_types": retrieval.get("rag_doc_types", ""),
                    "rag_titles": retrieval.get("rag_titles", ""),
                    "rag_model_response": rag_score.get("rag_model_response", ""),
                    "rag_score_reason": rag_score.get("rag_score_reason", ""),
                }
                resume_cache[case_row["case_id"]] = cache_entry_from_rows(progress_rag_row, retrieval)
                save_progress(output_dir, resume_cache, {
                    "case_count_after_dedupe": len(unique_cases),
                    "latest_case_id": case_row["case_id"],
                    "skipped_completed_count": skipped_completed_count,
                    "newly_executed_count": newly_executed_count,
                    "retried_failed_count": retried_failed_count,
                    "consecutive_service_failures": consecutive_service_failures,
                })
            if consecutive_service_failures >= 3:
                aborted_due_to_service_failure = True
                notes.append("检测到连续 3 次服务不可达，已停止正式续跑；已完成结果保存在 checkpoint 中。")
                print("连续 3 次服务不可达，停止续跑，避免将后续 case 误记为 0 分。", file=sys.stderr)
                break
            if args.sleep_ms:
                time.sleep(args.sleep_ms / 1000)

        rag_row = {
            "case_id": case_row["case_id"],
            "case_title": case_row["case_title"],
            "test_group": case_row["test_group"],
            "user_input": case_row["user_input"],
            **{field: rag_score.get(field, "") for field in SCORE_FIELDS},
            "rag_enabled": bool(args.rag_enabled and not args.no_rag_only),
            "rag_hit_count": retrieval.get("rag_hit_count", ""),
            "rag_doc_types": retrieval.get("rag_doc_types", ""),
            "rag_titles": retrieval.get("rag_titles", ""),
            "rag_request_time_ms": rag_score.get("rag_request_time_ms", retrieval.get("rag_request_time_ms", "")),
            "rag_failed_but_fallback": retrieval.get("rag_failed_but_fallback", ""),
            "rag_model_response": rag_score.get("rag_model_response", ""),
            "rag_score_reason": rag_score.get("rag_score_reason", ""),
        }
        rag_rows.append(rag_row)
        retrieval_rows.append(retrieval)

        no_total = no_rag.get("total_score")
        rag_total = rag_score.get("total_score")
        no_hit_rate = no_rag.get("must_ask_hit_rate")
        rag_hit_rate = rag_score.get("must_ask_hit_rate")
        score_delta = round(rag_total - no_total, 4) if isinstance(no_total, (int, float)) and isinstance(rag_total, (int, float)) else ""
        hit_delta = round(rag_hit_rate - no_hit_rate, 4) if isinstance(no_hit_rate, (int, float)) and isinstance(rag_hit_rate, (int, float)) else ""
        comparison_rows.append({
            "case_id": case_row["case_id"],
            "case_title": case_row["case_title"],
            "test_group": case_row["test_group"],
            "population_type": case_row["population_type"],
            "disease_category": case_row["disease_category"],
            "mode": case_row["mode"],
            "user_input": case_row["user_input"],
            "no_rag_total_score": no_total if no_total is not None else "",
            "rag_total_score": rag_total if rag_total is not None else "",
            "score_delta": score_delta,
            "no_rag_must_ask_hit_count": no_rag.get("must_ask_hit_count", ""),
            "rag_must_ask_hit_count": rag_score.get("must_ask_hit_count", ""),
            "must_ask_hit_count_delta": (rag_score.get("must_ask_hit_count", 0) - no_rag.get("must_ask_hit_count", 0)) if rag_score and no_rag else "",
            "no_rag_must_ask_hit_rate": no_hit_rate if no_hit_rate is not None else "",
            "rag_must_ask_hit_rate": rag_hit_rate if rag_hit_rate is not None else "",
            "must_ask_hit_rate_delta": hit_delta,
            "no_rag_medical_safety_score": no_rag.get("medical_safety_score", ""),
            "rag_medical_safety_score": rag_score.get("medical_safety_score", ""),
            "medical_safety_delta": (rag_score.get("medical_safety_score", 0) - no_rag.get("medical_safety_score", 0)) if rag_score and no_rag else "",
            "no_rag_department_accuracy_score": no_rag.get("department_accuracy_score", ""),
            "rag_department_accuracy_score": rag_score.get("department_accuracy_score", ""),
            "department_accuracy_delta": (rag_score.get("department_accuracy_score", 0) - no_rag.get("department_accuracy_score", 0)) if rag_score and no_rag else "",
            "no_rag_urgency_accuracy_score": "",
            "rag_urgency_accuracy_score": "",
            "urgency_accuracy_delta": "",
            "no_rag_is_one_vote_veto": no_rag.get("is_one_vote_veto", ""),
            "rag_is_one_vote_veto": rag_score.get("is_one_vote_veto", ""),
            "no_rag_is_suspected_failure": no_rag.get("is_suspected_failure", ""),
            "rag_is_suspected_failure": rag_score.get("is_suspected_failure", ""),
            "no_rag_request_error": no_rag.get("request_error", ""),
            "rag_request_error": rag_score.get("request_error", ""),
            "improved": score_delta != "" and score_delta > 0,
            "regressed": score_delta != "" and score_delta < 0,
            "unchanged": score_delta != "" and score_delta == 0,
            "no_rag_model_response": no_rag.get("model_response", ""),
            "rag_model_response": rag_score.get("rag_model_response", ""),
            "comparison_note": "无历史无 RAG 评分" if not no_rag else ("no-rag-only 未运行 RAG" if args.no_rag_only else ""),
        })

    if aborted_due_to_service_failure:
        progress_path = output_dir / PROGRESS_FILE
        print(f"已保存断点进度: {progress_path}", file=sys.stderr)
        return 6

    low_score_cases = [
        row for row in comparison_rows
        if row.get("rag_request_error") or row.get("rag_is_one_vote_veto") or row.get("rag_is_suspected_failure")
        or (isinstance(row.get("rag_total_score"), (int, float)) and row["rag_total_score"] < 8)
        or (isinstance(row.get("rag_must_ask_hit_rate"), (int, float)) and row["rag_must_ask_hit_rate"] < 0.4)
    ]
    regressed_cases = [
        row for row in comparison_rows
        if row.get("regressed") or (
            isinstance(row.get("must_ask_hit_rate_delta"), (int, float)) and row["must_ask_hit_rate_delta"] < 0
        )
    ]
    summary_by_group = build_summaries(comparison_rows, "test_group")
    summary_by_population = build_summaries(comparison_rows, "population_type")
    summary_by_disease = build_summaries(comparison_rows, "disease_category")

    summary = {
        "history_file_count": len(history_files),
        "case_count_before_dedupe": len(case_rows),
        "case_count_after_dedupe": len(unique_cases),
        "duplicate_count": len(duplicates),
        "no_rag_score_missing_count": sum(1 for row in unique_cases if row["no_rag_score_missing"]),
        "rag_actual_test_count": sum(
            1
            for row in rag_rows
            if row.get("rag_model_response") or row.get("total_score") not in (None, "")
        ),
        "rag_request_error_count": sum(1 for row in rag_rows if row.get("request_error")),
        "no_rag_avg_total_score": mean([row.get("total_score") for row in no_rag_rows]),
        "rag_avg_total_score": mean([row.get("total_score") for row in rag_rows]),
        "avg_score_delta": mean([row.get("score_delta") for row in comparison_rows]),
        "no_rag_avg_must_ask_hit_rate": mean([row.get("must_ask_hit_rate") for row in no_rag_rows]),
        "rag_avg_must_ask_hit_rate": mean([row.get("must_ask_hit_rate") for row in rag_rows]),
        "avg_must_ask_hit_rate_delta": mean([row.get("must_ask_hit_rate_delta") for row in comparison_rows]),
        "improved_count": sum(1 for row in comparison_rows if row.get("improved")),
        "regressed_count": sum(1 for row in comparison_rows if row.get("regressed")),
        "unchanged_count": sum(1 for row in comparison_rows if row.get("unchanged")),
        "skipped_completed_count": skipped_completed_count,
        "newly_executed_count": newly_executed_count,
        "retried_failed_count": retried_failed_count,
        "resume_source": resume_source,
        "aborted_due_to_service_failure": aborted_due_to_service_failure,
        **env,
    }

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme_rows = [
        {"item": "测试目的", "value": "整理历史无 RAG 测试集，并用同一批 case 对当前 RAG 模型做 A/B 对比。"},
        {"item": "数据来源", "value": str(reports_dir)},
        {"item": "去重规则", "value": "优先 case_id；缺失时使用 user_input/test_group/case_title/expected_department/expected_urgency；同文本视为重复。"},
        {"item": "测试时间", "value": timestamp},
        {"item": "同一批 case", "value": "是"},
        {"item": "评分维度一致", "value": "是，复用 evaluation/scripts 现有评分逻辑。"},
        {"item": "RAG 配置状态", "value": summary.get("rag_enabled_status")},
        {"item": "Milvus collection 状态", "value": stringify({k: summary.get(k) for k in ["collection_exists", "num_entities", "doc_type_counts"]})},
        {"item": "注意事项", "value": "; ".join(notes) if notes else "正式 RAG 测试结果来自当前后端接口。"},
    ]
    readme.append(["item", "value"])
    for row in readme_rows:
        readme.append([row["item"], row["value"]])
    style_sheet(readme)

    add_sheet(workbook, "history_files", history_files, ["path", "file_type", "sheets", "rows", "columns", "purpose"])
    add_sheet(workbook, "unique_test_cases", [{k: v for k, v in row.items() if not k.startswith("_")} for row in unique_cases], CASE_FIELDS)
    add_sheet(workbook, "duplicates", duplicates, ["kept_case_id", "duplicate_case_id", "kept_source", "duplicate_source", "duplicate_title"])
    add_sheet(workbook, "no_rag_scores", no_rag_rows, list(no_rag_rows[0].keys()) if no_rag_rows else ["case_id"])
    add_sheet(workbook, "rag_scores", rag_rows, list(rag_rows[0].keys()) if rag_rows else ["case_id"])
    add_sheet(workbook, "comparison", comparison_rows, list(comparison_rows[0].keys()) if comparison_rows else ["case_id"])
    add_sheet(workbook, "summary_by_group", summary_by_group, list(summary_by_group[0].keys()) if summary_by_group else ["test_group"])
    add_sheet(workbook, "summary_by_population", summary_by_population, list(summary_by_population[0].keys()) if summary_by_population else ["population_type"])
    add_sheet(workbook, "summary_by_disease_category", summary_by_disease, list(summary_by_disease[0].keys()) if summary_by_disease else ["disease_category"])
    add_sheet(workbook, "low_score_cases", low_score_cases, list(comparison_rows[0].keys()) if comparison_rows else ["case_id"])
    add_sheet(workbook, "regressed_cases", regressed_cases, list(comparison_rows[0].keys()) if comparison_rows else ["case_id"])
    add_sheet(workbook, "rag_retrieval_details", retrieval_rows, list(retrieval_rows[0].keys()) if retrieval_rows else ["case_id"])

    workbook.save(output_xlsx)
    write_report_md(output_md, summary, summary_by_group, notes)

    print("历史测试文件数量:", summary["history_file_count"])
    print("去重前 case 数:", summary["case_count_before_dedupe"])
    print("去重后唯一 case 数:", summary["case_count_after_dedupe"])
    print("重复 case 数:", summary["duplicate_count"])
    print("跳过已完成 case 数:", summary["skipped_completed_count"])
    print("本次新执行 case 数:", summary["newly_executed_count"])
    print("重试失败 case 数:", summary["retried_failed_count"])
    print("有 RAG 实际测试 case 数:", summary["rag_actual_test_count"])
    print("请求失败 case 数:", summary["rag_request_error_count"])
    print("无 RAG 平均分:", summary["no_rag_avg_total_score"])
    print("有 RAG 平均分:", summary["rag_avg_total_score"])
    print("平均分变化:", summary["avg_score_delta"])
    print("无 RAG 必问命中率:", summary["no_rag_avg_must_ask_hit_rate"])
    print("有 RAG 必问命中率:", summary["rag_avg_must_ask_hit_rate"])
    print("必问命中率变化:", summary["avg_must_ask_hit_rate_delta"])
    print("输出 Excel:", output_xlsx)
    print("输出 Markdown:", output_md)
    print("需人工复核 case:", len(low_score_cases))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
