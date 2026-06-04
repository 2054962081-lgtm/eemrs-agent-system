#!/usr/bin/env python3
"""Export the extended Excel test sheet to the JSONL format used by eval runner."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "medical_agent_test_cases_extended.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "medical_agent_test_cases_extended_special.jsonl"
DEFAULT_SHEET = "新增专项测试样例"


def split_items(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in text.replace("；", ";").split(";") if item.strip()]


def normalize_urgency(value: Any) -> str:
    text = str(value or "").strip()
    if any(token in text for token in ("120", "立即", "急诊")):
        return "emergency"
    if any(token in text for token in ("尽快", "及时", "当天")):
        return "urgent"
    return "normal"


def row_to_case(row: dict[str, Any]) -> dict[str, Any]:
    user_input = str(row.get("user_input") or "").strip()
    departments = split_items(row.get("expected_department") or row.get("expectedDepartments"))
    must_ask = split_items(row.get("must_ask_points") or row.get("mustAskOrMention"))
    forbidden = split_items(row.get("forbidden_actions") or row.get("mustNotSay"))
    red_flags = split_items(row.get("red_flags"))

    return {
        "caseId": str(row.get("caseId") or "").strip(),
        "caseType": str(row.get("caseType") or row.get("test_group") or "extended_special").strip(),
        "mode": str(row.get("mode") or "quick").strip(),
        "title": str(row.get("title") or "").strip(),
        "patientProfile": {
            "age": row.get("patientAge"),
            "gender": str(row.get("patientGender") or "").strip(),
            "specialPopulation": str(row.get("population_type") or row.get("specialPopulation") or "").strip(),
            "chronicDisease": "",
        },
        "turns": [{"role": "user", "content": user_input}],
        "expected": {
            "recommendedDepartments": departments,
            "urgency": normalize_urgency(row.get("expected_urgency") or row.get("expectedUrgency")),
            "mustAskOrMention": must_ask,
            "mustNotSay": forbidden,
            "redFlagsToMention": red_flags,
        },
        "scoring": {
            "medicalSafety": 3,
            "departmentAccuracy": 2,
            "consultationCompleteness": 2,
            "modeCompliance": 1,
            "antiMisleading": 1,
            "userExperience": 1,
        },
        "hardFailRules": split_items(row.get("hardFailRules")) or forbidden,
        "notes": str(row.get("notes") or row.get("scoring_focus") or "").strip(),
    }


def read_sheet(input_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [str(cell or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(value is not None and str(value).strip() for value in row.values()):
            rows.append(row)
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    rows = read_sheet(Path(args.input), args.sheet)
    cases = [row_to_case(row) for row in rows]
    missing_input = [case["caseId"] for case in cases if not case["turns"][0]["content"]]
    if missing_input:
        raise ValueError(f"Cases missing user input: {', '.join(missing_input)}")

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        for case in cases:
            file.write(json.dumps(case, ensure_ascii=False) + "\n")

    group_counts: dict[str, int] = {}
    for row in rows:
        group = str(row.get("test_group") or "未分组")
        group_counts[group] = group_counts.get(group, 0) + 1

    print(f"Exported cases: {len(cases)}")
    print(f"Output JSONL: {output_path}")
    print("Group counts:")
    for group, count in group_counts.items():
        print(f"- {group}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
