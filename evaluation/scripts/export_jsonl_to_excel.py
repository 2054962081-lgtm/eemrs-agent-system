#!/usr/bin/env python3
"""Export the JSONL evaluation set to a reviewer-friendly Excel file."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - depends on local environment
    raise SystemExit("openpyxl is required to export Excel: pip install openpyxl") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "pre_consultation_eval_cases.jsonl"
DEFAULT_OUTPUT = ROOT / "data" / "pre_consultation_eval_cases.sample.xlsx"

HEADERS = [
    "caseId",
    "caseType",
    "mode",
    "title",
    "patientAge",
    "patientGender",
    "specialPopulation",
    "turnsText",
    "expectedDepartments",
    "expectedUrgency",
    "mustAskOrMention",
    "mustNotSay",
    "hardFailRules",
    "notes",
]


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    cases: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, 1):
            stripped = line.strip()
            if not stripped:
                continue
            try:
                cases.append(json.loads(stripped))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON on line {line_number}: {exc}") from exc
    return cases


def join_values(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def flatten_case(case: dict[str, Any]) -> dict[str, str]:
    profile = case.get("patientProfile") or {}
    expected = case.get("expected") or case.get("expectedRecord") or {}
    turns = case.get("turns") or case.get("history") or []
    turns_text = "\n".join(
        f"{turn.get('role', '')}: {turn.get('content', '')}" for turn in turns
    )
    departments = (
        expected.get("recommendedDepartments")
        or expected.get("departments")
        or expected.get("department")
        or []
    )
    return {
        "caseId": join_values(case.get("caseId")),
        "caseType": join_values(case.get("caseType")),
        "mode": join_values(case.get("mode")),
        "title": join_values(case.get("title")),
        "patientAge": join_values(profile.get("age")),
        "patientGender": join_values(profile.get("gender")),
        "specialPopulation": join_values(profile.get("specialPopulation")),
        "turnsText": turns_text,
        "expectedDepartments": join_values(departments),
        "expectedUrgency": join_values(expected.get("urgency")),
        "mustAskOrMention": join_values(expected.get("mustAskOrMention")),
        "mustNotSay": join_values(expected.get("mustNotSay") or expected.get("mustNotInvent")),
        "hardFailRules": join_values(case.get("hardFailRules")),
        "notes": join_values(case.get("notes")),
    }


def export_excel(cases: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "pre_consultation_eval"
    sheet.append(HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for case in cases:
        row = flatten_case(case)
        sheet.append([row[header] for header in HEADERS])

    widths = {
        "A": 14,
        "B": 20,
        "C": 10,
        "D": 28,
        "E": 12,
        "F": 14,
        "G": 22,
        "H": 60,
        "I": 28,
        "J": 16,
        "K": 48,
        "L": 42,
        "M": 50,
        "N": 40,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{sheet.max_row}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="JSONL input path")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="XLSX output path")
    args = parser.parse_args()

    cases = read_jsonl(Path(args.input))
    export_excel(cases, Path(args.output))
    print(f"Exported {len(cases)} cases to {args.output}")


if __name__ == "__main__":
    main()
